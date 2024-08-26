from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Path to the Excel file
EXCEL_FILE = 'data.xlsx'

def append_to_excel(data):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Item', 'Status'])
    
    for item, status in data.items():
        ws.append([item, status])
    
    wb.save(EXCEL_FILE)

@app.route('/submit', methods=['POST'])
def submit():
    try:
        data = request.json
        append_to_excel(data)
        return jsonify({'message': 'Data saved successfully!'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
